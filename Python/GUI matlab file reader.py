import os
import io
import csv
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
from openpyxl.utils import get_column_letter

def load_sim_csv(path):
    lines = open(path, 'r', encoding='utf-8').readlines()
    # find the duplicated header line
    header_i = None
    for i in range(len(lines)-1):
        if lines[i].strip() and lines[i] == lines[i+1]:
            header_i = i
            break
    if header_i is None:
        raise ValueError("Could not locate duplicated header lines in CSV.")
    header = lines[header_i].strip()
    data_start = header_i + 4  # header, dup header, units row, blank -> +4
    csv_text = header + '\n' + ''.join(lines[data_start:])
    return pd.read_csv(io.StringIO(csv_text))

# ---- metrics ----
func_map = {
    
    'lap_time':                 lambda df: df['time'].max(),
    
    'distance':                 lambda df: df['distance'].max() - df['distance'].min(),
    
    'max_speed':                lambda df: df['speed'].max() * 3.6,
    'average_speed':            lambda df: df['speed'].mean() * 3.6,
    
    'energy_spent':             lambda df: df['energy_spent_fuel'].max(),
    'energy_spent_mech_max':    lambda df: df['energy_spent_mech'].max(),
    
    'ax_min':                   lambda df: df['long_acc'].min(),
    'ax_max':                   lambda df: df['long_acc'].max(),
    'ax_mean':                  lambda df: df['long_acc'].mean(),
    'ay_max':                   lambda df: df['lat_acc'].max(),
    'ay_mean':                  lambda df: df['lat_acc'].mean(),
    
    'wheel_torque_max':         lambda df: df['wheel_torque'].max(),
    'wheel_torque_mean':        lambda df: df['wheel_torque'].mean(),
    'engine_torque_max':        lambda df: df['engine_torque'].max(),
    'engine_torque_mean':       lambda df: df['engine_torque'].mean(),
    
    'engine_power_max':         lambda df: df['engine_power'].max(),
    'engine_power_mean':        lambda df: df['engine_power'].mean(),
    'engine_speed_max':         lambda df: df['engine_speed'].max(),
    'engine_speed_average':     lambda df: df['engine_speed'].mean(),
    
    'capacity used'
    
    'throttle_mean':            lambda df: df['throttle'].mean(),
    'brake_max':                lambda df: df['brake_force'].max(),
    'brake_mean':               lambda df: df['brake_force'].mean(),
    
    'Fz_aero_max':              lambda df: -df['Fz_aero'].min(),
    'Fz_aero_mean':             lambda df: -df['Fz_aero'].mean(),
    'Fx_aero_max':              lambda df: -df['Fx_aero'].min(),
    'Fx_aero_mean':             lambda df: -df['Fx_aero'].mean(),
    
    'Fx_eng_max':               lambda df: df['Fx_eng'].max(),
    'Fx_eng_average':           lambda df: df['Fx_eng'].mean(),
    'Fx_roll_max':              lambda df: -df['Fx_roll'].min(),
    'Fx_roll_average':          lambda df: -df['Fx_roll'].mean(),
    
    'Fz_mass_max':              lambda df: -df['Fz_mass'].min(),
    'Fz_mass_average':          lambda df: -df['Fz_mass'].mean(),
    'Fz_total_max':             lambda df: -df['Fz_total'].min(),
    'Fz_total_mean':            lambda df: -df['Fz_total'].mean(),
    
    'elevation_max':            lambda df: df['elevation'].max(),
    'elevation_average':        lambda df: df['elevation'].mean(),
    
    'yaw_rate_max':             lambda df: df['yaw_rate'].max(),
    'yaw_rate_average':         lambda df: df['yaw_rate'].mean(),
    
    'brake_pres_max':           lambda df: df['brake_pres'].max(),
    'brake_pres_average':       lambda df: df['brake_pres'].mean(),
    
    'steering_max':             lambda df: df['steering'].max(),
    'steering_average':         lambda df: df['steering'].mean(),
    'delta_max':                lambda df: df['delta'].max(),
    'delta_average':            lambda df: df['delta'].mean(),
    'beta_max':                 lambda df: df['beta'].max(),
    'beta_average':             lambda df: df['beta'].mean(),
}

unit_map = {
    'lap_time':'s','distance':'m','max_speed':'km/h','average_speed':'km/h',
    'energy_spent':'J','energy_spent_mech_max':'J',
    'ax_min':'m/s²','ax_max':'m/s²','ax_mean':'m/s²',
    'ay_max':'m/s²','ay_mean':'m/s²',
    'wheel_torque_max':'Nm','wheel_torque_mean':'Nm',
    'engine_torque_max':'Nm','engine_torque_mean':'Nm',
    'engine_power_max':'W','engine_power_mean':'W',
    'engine_speed_max':'rpm','engine_speed_average':'rpm',
    'throttle_mean':'%','brake_max':'N','brake_mean':'N',
    'Fz_aero_max':'N','Fz_aero_mean':'N','Fx_aero_max':'N','Fx_aero_mean':'N',
    'Fx_eng_max':'N','Fx_eng_average':'N','Fx_roll_max':'N','Fx_roll_average':'N',
    'Fz_mass_max':'N','Fz_mass_average':'N',
    'Fz_total_max':'N','Fz_total_mean':'N',
    'elevation_max':'m','elevation_average':'m',
    'yaw_rate_max':'°/s','yaw_rate_average':'°/s',
    'brake_pres_max':'bar','brake_pres_average':'bar',
    'steering_max':'°','steering_average':'°',
    'delta_max':'','delta_average':'','beta_max':'','beta_average':'',
    'FS Event Score':'points',
    'simulation':''  # custom field
}

# ---- FS event scoring ----
def acceleration_score(t_team: float) -> float:
    P_max, T_max = 50, 1.5 * 3.454
    t = min(t_team, T_max)
    raw = ((T_max / t) - 1) / 0.5
    res = 0.95 * P_max * raw + 0.05 * P_max
    ret = min(P_max, res)
    return ret

def skidpad_score(t_team: float) -> float:
    t_team = t_team/2
    
    P_max = 50 
    T_max = 1.25 * 4.627 #2024 fastest time FSG
    t = min(t_team, T_max)
    raw = ((((T_max / t) ** 2) - 1) / 0.5625)
    res = 0.95 * P_max * raw + 0.05 * P_max
    ret = min(P_max, res)
    return ret

def autocross_score(t_team: float) -> float:
    P_max, T_max = 100, 1.25 * 75.93
    t = min(t_team, T_max)
    frac = (T_max / t) - 1
    term = frac / 0.25 if frac > 0 else 0
    res = 0.95 * P_max * term + 0.05 * P_max
    ret = ret = min(P_max, res)
    return ret

def endurance_score(t_lap: float) -> float:
    P_max, T_max = 250, 1.333 * 1420.56
    T_team = 18 * t_lap
    t = min(T_team, T_max)
    raw = ((T_max / t) - 1) / 0.333
    res = 0.90 * P_max * raw + 0.10 * P_max
    ret = min(P_max, res)
    
    return ret

event_map = {
    'NO FS EVENT':   None,
    'Acceleration':  acceleration_score,
    'Skidpad':       skidpad_score,
    'Autocross':     autocross_score,
    'Endurance':     endurance_score,
}

defaults = {
    'simulation','FS Event Score','lap_time','distance','max_speed','energy_spent',
    'ax_min','ax_max','ay_max','engine_torque_max','engine_power_max',
    'throttle_mean','Fz_aero_max','Fx_aero_max','engine_speed_max','engine_speed_average'
}

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Simulation Summary")
        w = int(self.winfo_screenwidth()*0.75)
        h = int(self.winfo_screenheight()*0.75)
        self.geometry(f"{w}x{h}+0+0")

        # Top controls
        top = ttk.Frame(self); top.pack(fill='x', padx=10, pady=5)
        ttk.Label(top, text="CSV:").pack(side='left')
        self.csv_path = tk.StringVar()
        ttk.Entry(top, textvariable=self.csv_path, width=40).pack(side='left', padx=5)
        ttk.Button(top, text="Browse…", command=self.browse_csv).pack(side='left', padx=(0,20))
        ttk.Label(top, text="FS Event:").pack(side='left')
        self.event_var = tk.StringVar(value='NO FS EVENT')
        ttk.Combobox(top, textvariable=self.event_var,
                     values=list(event_map.keys()), state='readonly', width=15).pack(side='left', padx=(0,20))
        ttk.Label(top, text="Output XLSX:").pack(side='left')
        self.out_path = tk.StringVar()
        ttk.Entry(top, textvariable=self.out_path, width=40).pack(side='left', padx=5)
        ttk.Button(top, text="Save As…", command=self.save_as).pack(side='left', padx=(0,20))
        ttk.Label(top, text="Simulation:").pack(side='left')
        self.simulation = tk.StringVar()
        ttk.Entry(top, textvariable=self.simulation, width=20).pack(side='left', padx=(0,20))
        ttk.Button(top, text="Compute", command=self.compute).pack(side='left', padx=5)
        ttk.Button(top, text="Export",  command=self.export).pack(side='left', padx=5)
        ttk.Button(top, text="Exit",    command=self.destroy).pack(side='left', padx=5)

        # Main panes
        main = ttk.Frame(self); main.pack(fill='both', expand=True, padx=10, pady=5)
        main.rowconfigure(0, weight=1)
        main.columnconfigure(0, weight=0)
        main.columnconfigure(2, weight=1)

        # Left: scrollable checkboxes
        left_frame = ttk.LabelFrame(main, text="Show in GUI", width=250)
        left_frame.grid(row=0, column=0, sticky='nsew', padx=(0,10), pady=5)
        canvas = tk.Canvas(left_frame)
        scroll = ttk.Scrollbar(left_frame, orient='vertical', command=canvas.yview)
        box_frame = ttk.Frame(canvas)
        box_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0,0), window=box_frame, anchor='nw')
        canvas.configure(yscrollcommand=scroll.set)
        canvas.pack(side='left', fill='both', expand=True)
        scroll.pack(side='right', fill='y')

        self.chk_vars = {}
        for name in ['simulation','FS Event Score'] + list(func_map):
            var = tk.BooleanVar(value=(name in defaults))
            ttk.Checkbutton(box_frame, text=name, variable=var).pack(anchor='w', pady=2, padx=5)
            self.chk_vars[name] = var

        # Right: values table
        right = ttk.Frame(main); right.grid(row=0, column=2, sticky='nsew', pady=4)
        self.tree = ttk.Treeview(right, columns=('Metric','Value','Unit'), show='headings')
        for col, width in [('Metric',200), ('Value',120), ('Unit',70)]:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=width, anchor='e' if col=='Value' else 'w', stretch=False)
        self.tree.pack(fill='both', expand=True)

        self.df = None
        self.summary = {}
        self.raw_lines = []

    def browse_csv(self):
        p = filedialog.askopenfilename(filetypes=[("CSV","*.csv")])
        if p: self.csv_path.set(p)

    def save_as(self):
        p = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                         filetypes=[("Excel","*.xlsx")])
        if p: self.out_path.set(p)

    def compute(self):
        if not self.csv_path.get():
            messagebox.showerror("Error","Select a CSV first"); return
        try:
            self.df = load_sim_csv(self.csv_path.get())
            with open(self.csv_path.get(), 'r', encoding='utf-8') as f:
                self.raw_lines = [ln.rstrip('\n') for ln in f]
        except Exception as e:
            messagebox.showerror("Error", f"Cannot read CSV:\n{e}"); return

        # compute metrics
        self.summary = {k: round(f(self.df),2) for k,f in func_map.items()}
        # FS event score
        ev = self.event_var.get()
        if ev != 'NO FS EVENT':
            t = self.summary['lap_time']
            self.summary['FS Event Score'] = round(event_map[ev](t),2)
        # Simulation label
        sim_val = (self.simulation.get().strip()
                   or os.path.splitext(os.path.basename(self.csv_path.get()))[0])
        self.summary['simulation'] = sim_val

        # refresh table
        for i in self.tree.get_children(): self.tree.delete(i)
        for name, var in self.chk_vars.items():
            if var.get() and name in self.summary:
                self.tree.insert('', 'end', values=(name, self.summary[name], unit_map.get(name,'')))

    def export(self):
        if not self.summary or self.df is None:
            messagebox.showerror("Error","Compute first"); return
        if not self.out_path.get():
            messagebox.showerror("Error","Choose output XLSX"); return

        # put 'simulation' first
        keys = ['simulation'] + [k for k in self.summary.keys() if k != 'simulation']
        df_out = pd.DataFrame({
            'Metric': keys,
            'Value':  [self.summary[k] for k in keys],
            'Unit':   [unit_map.get(k,'') for k in keys]
        })

        # Expand entire CSV into grid
        raw_rows = list(csv.reader(self.raw_lines))
        max_cols = max((len(r) for r in raw_rows), default=1)

        # find header row (duplicated header lines)
        header_i = None
        for i in range(len(self.raw_lines)-1):
            if self.raw_lines[i].strip() and self.raw_lines[i] == self.raw_lines[i+1]:
                header_i = i
                break

        try:
            outdir = os.path.dirname(self.out_path.get())
            if outdir:
                os.makedirs(outdir, exist_ok=True)
            with pd.ExcelWriter(self.out_path.get(), engine='openpyxl', mode='w') as w:
                # summary sheet
                df_out.to_excel(w, index=False, sheet_name='summary')

                # data sheet as clickable grid
                wb = w.book
                ws = wb.create_sheet('data')
                for row in raw_rows:
                    ws.append(row)

                # freeze panes & filter starting from real header
                if header_i is not None:
                    ws.freeze_panes = f"A{header_i+2}"
                    last_col_letter = get_column_letter(max_cols)
                    last_row = len(raw_rows)
                    ws.auto_filter.ref = f"A{header_i+1}:{last_col_letter}{last_row}"

            messagebox.showinfo("Done", f"Wrote summary and full CSV grid to\n{self.out_path.get()}")
        except Exception as e:
            messagebox.showerror("Error", f"Write failed:\n{e}")

if __name__ == "__main__":
    App().mainloop()
